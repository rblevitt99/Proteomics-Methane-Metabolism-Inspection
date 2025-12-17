# The goal of all this code is to take all KEGG identifiers associated with methane metabolism and 
# merge them to my protoemics dataframe. This will tell me how methane metabolism changes across all conditions. 

from Bio import SeqIO
import pandas as pd
import openpyxl
import xlsxwriter


# Data = "Merged_Protein_and_KEGG_Data.xlsx"


# Here is a quick function to grab the columns I want to work with from Protein)_data
 
workbook=openpyxl.Workbook()
workbook.save(filename="Methane_Metabolism_Proteomics.xlsx")

sheet=["LowCa_vs_Ca","Ca_vs_Low_Fe","LowFe_vs_Fe","Fe_vs_Mix","Mix_vs_Ni","Ni_vs_Nd","Nd_vs_W","W_vs_Cu"]

print(sheet)

def Methane_Grabber(input):
    Protein_Data = pd.read_excel(Data, sheet_name=input)
    Desired_Protein_Data = Protein_Data[['Averaged_Log2Ratio','p-value','locus_tag','refseq_protein_id_x','KEGG_Info']]
    Methane = pd.read_excel("Methane_Metabolism_KEGG_Identifiers.xlsx")
    with pd.ExcelWriter("Methane_Metabolism_Proteomics.xlsx", mode="a", engine="openpyxl", if_sheet_exists="new") as writer:
        merged_df = pd.merge(Desired_Protein_Data,Methane,on='KEGG_Info', how='right') # Joining the proteomic and KEGG data on the refseq_protein_id. Inner will return rows that only existed in both data frames.
        merged_df.to_excel(writer, sheet_name=input, index=False)    

for input in sheet:
    Methane_Grabber(input)


sheet=["LowCa_vs_Ca","Ca_vs_Low_Fe", "LowFe_vs_Fe","Fe_vs_Mix","Mix_vs_Ni",
"Ni_vs_Nd","Nd_vs_W","W_vs_Cu"]

def Methane_Grabber(input):
    Protein_Data = pd.read_excel(Data, sheet_name=input)
    Desired_Protein_Data = Protein_Data[['Averaged_Log2Ratio','locus_tag','refseq_protein_id_x','KEGG_Info']]
    Methane= pd.read_excel("Methane_Metabolism_KEGG_Identifiers.xlsx")
    merged_df = pd.merge(Desired_Protein_Data,Methane,on='KEGG_Info', how='right') # Joining the proteomic and KEGG data on the refseq_protein_id. Inner will return rows that only existed in both data frames.

for input in sheet:
    Methane_Grabber(input)


Meth = "Methane_Metabolism_Proteomics.xlsx"

# Prot_meth = pd.read_excel(Meth, sheet_name = "LowCa_vs_Ca")

#print(Prot_meth)

#KEGG_LowCa = Prot_meth['KEGG_Info']

#print(f"Condition: Condition - {input}", len(KEGG))

# def Counter(input):
#     Prot_meth = pd.read_excel(Meth, sheet_name = input)
#     KEGG = Prot_meth['KEGG_Info']
#     print(f"Condition:{input}", len(KEGG), KEGG)

# for input in sheet:
#     Counter(input)
    

# workbook=openpyxl.Workbook()
# workbook.save(filename="Methane_Metabolism_Proteomics_KEGG.xlsx")

# def Grabber(input):
#     Meth_Protein_Data = pd.read_excel(Meth, sheet_name=input)
#     Reference= pd.read_excel("Low_Ca_Kegg.xlsx")
#     with pd.ExcelWriter("Methane_Metabolism_Proteomics_KEGG.xlsx", mode="a", engine="openpyxl", if_sheet_exists="new") as writer:
#         merged_df = pd.merge(Meth_Protein_Data,Reference,on='KEGG_Info', how='left') 
#         merged_df= merged_df[['Averaged_Log2Ratio_x','locus_tag_x','p-value_x','refseq_protein_id_x_x','KEGG_Info']]
#         merged_df.to_excel(writer, sheet_name=input, index=False)

# for input in sheet:
#     Grabber(input)


# locus_id_map_WP = {} 

# with open("plasmid__genome_20z_WP") as input_handle:
#       for record in SeqIO.parse(input_handle, "genbank"):
#         #print(record)



#         for feature in record.features:
#             curr_feature_dict = feature.qualifiers
#             #print(curr_feature_dict)
#             if 'old_locus_tag' not in curr_feature_dict: 
#                 continue
#             if 'protein_id' not in curr_feature_dict:
#                 continue
#             MEALZ_Value = curr_feature_dict['old_locus_tag'] # Grabs the value associated with locus_tag
#             Protein_ID = curr_feature_dict['protein_id'] # Grabs the value associated with protein_id
#             string_MZ = MEALZ_Value[0] #extracts string from the list
#             #print(string_MZ)
#             string_Prot = Protein_ID[0] #extracts string from the list
#            # print(string_Prot)
#             locus_id_map_WP[string_Prot] = string_MZ # Makes protein_id the key and MEALZ_Value the value. All being added to locus_id_map_WP.
# #print(locus_id_map_WP)

# This code below is going into the WP 20z genome file and picking out the protein id and product. All while putting it into a dictionary

# protein_and_product= {} 

#print(type(curr_feature_dict))


# with open("plasmid__genome_20z_WP") as input_handle:
#     for record in SeqIO.parse(input_handle, "genbank"):
#         with open('20z_Genome_features.txt', 'w') as file:  # WRONG: opens each loop
#             for feature in record.features:
#                 curr_feature_dict = feature.qualifiers
#                 file.write(str(curr_feature_dict) + "\n\n")





with open("plasmid__genome_20z_WP") as input_handle:
      for record in SeqIO.parse(input_handle, "genbank"):
        for feature in record.features:
            curr_feature_dict = feature.qualifiers
            print(curr_feature_dict)

# with open("plasmid__genome_20z_WP") as input_handle:
#       for record in SeqIO.parse(input_handle, "genbank"):
#         #print(record)



#         for feature in record.features:
#             curr_feature_dict = feature.qualifiers
#             #print(curr_feature_dict)
#             if 'product' not in curr_feature_dict: 
#                 continue
#             if 'protein_id' not in curr_feature_dict:
#                 continue
#             Protein_product = curr_feature_dict['product'] # Grabs the value associated with product
#             Protein_ID = curr_feature_dict['protein_id'] # Grabs the value associated with protein_id
#             string_product = Protein_product[0] #extracts string from the list
#             string_Prot = Protein_ID[0] #extracts string from the list
#             protein_and_product[string_Prot] = string_product # Makes protein_id the key and product the value. All being added to protein_and_product .
# print(protein_and_product)

#This code below is taking the information in the dictionary: protein_and_product and putting it into an excel file

#df = pd.DataFrame(data=protein_and_product, index=[0]) # 

#df = (df.T)

#print(df)

#df.to_excel('Refseqprotein_and_product.xlsx')


# workbook=openpyxl.Workbook()
# workbook.save(filename="Proteomics_Methane_Metabolism_P_Products.xlsx")

# def Product_Merger(input):
#     Meth_Protein_Data = pd.read_excel(Meth, sheet_name=input)
#     Product= pd.read_excel("Refseqprotein_and_product.xlsx")
#     with pd.ExcelWriter("Proteomics_Methane_Metabolism_P_Products.xlsx", mode="a", engine="openpyxl", if_sheet_exists="new") as writer:
#         merged_df = pd.merge(Meth_Protein_Data,Product,on='protein_id', how='left') 
#         merged_df.to_excel(writer, sheet_name=input, index=False)

# for input in sheet:
#     Product_Merger(input)


# workbook=openpyxl.Workbook()
# workbook.save(filename="Proteomics_Methane_Metabolism_Lowfe_vs_fe.xlsx")


# Prot_Data = pd.read_excel(Meth,sheet_name='LowFe_vs_Fe')
# print(Prot_Data)



# Here is some code that will filter for DE protiens


# workbook=openpyxl.Workbook()
# workbook.save(filename="DE_Proteomics_Methane_Metabolism.xlsx")


# def DE(input):
#     Proteomic_data = pd.read_excel("Proteomics_Methane_Metabolism_P_Products.xlsx", sheet_name= input)
#     #DE = Proteomic_data[(Proteomic_data['Averaged_Log2Ratio'] > 1) | (Proteomic_data['Averaged_Log2Ratio'] < -1) & (Proteomic_data['p-value'] < 0.05)]
#     with pd.ExcelWriter("DE_Proteomics_Methane_Metabolism.xlsx", mode="a", engine="openpyxl", if_sheet_exists="new") as writer:
#         DE = Proteomic_data[((Proteomic_data['Averaged_Log2Ratio'] > 1) | (Proteomic_data['Averaged_Log2Ratio'] < -1)) & (Proteomic_data['p-value'] < 0.05)]
#         DE.to_excel(writer,sheet_name= input, index = False)

# for input in  sheet:
#     DE(input)